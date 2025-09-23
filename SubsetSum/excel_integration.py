#!/usr/bin/env python3
"""
Excel Integration: Read and write Excel files for FuzzySum.

This module provides utilities for reading numeric data from Excel files,
writing results back to Excel with formatting, and handling named ranges.
"""

import logging
from typing import List, Optional, Tuple, Dict, Any, Union
from pathlib import Path

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

from fuzzy_sum import A1Converter, SolverResult


class ExcelIntegrationError(Exception):
    """Base exception for Excel integration errors."""
    pass


class ExcelReader:
    """Read data from Excel files."""

    def __init__(self, file_path: Path):
        """Initialize Excel reader."""
        self.file_path = Path(file_path)
        self.logger = logging.getLogger(__name__)

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        # Check file extension
        ext = self.file_path.suffix.lower()
        if ext not in ['.xlsx', '.xls', '.xlsm']:
            raise ValueError(f"Unsupported Excel format: {ext}")

        self.is_xlsx = ext in ['.xlsx', '.xlsm']

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names in the workbook."""
        if HAS_PANDAS:
            try:
                with pd.ExcelFile(self.file_path) as xls:
                    return xls.sheet_names
            except Exception as e:
                self.logger.warning(f"Pandas failed to read sheet names: {e}")

        if HAS_OPENPYXL and self.is_xlsx:
            try:
                wb = openpyxl.load_workbook(self.file_path, read_only=True)
                return wb.sheetnames
            except Exception as e:
                self.logger.warning(f"openpyxl failed to read sheet names: {e}")

        if HAS_XLRD and not self.is_xlsx:
            try:
                with xlrd.open_workbook(self.file_path) as wb:
                    return wb.sheet_names()
            except Exception as e:
                self.logger.warning(f"xlrd failed to read sheet names: {e}")

        raise ExcelIntegrationError("No suitable Excel library available")

    def read_range(
        self,
        sheet_name: Optional[str] = None,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        column_only: bool = True
    ) -> List[float]:
        """
        Read numeric values from Excel range.

        Args:
            sheet_name: Sheet name (None for first sheet)
            start_cell: Starting cell address (e.g., "A1")
            end_cell: Ending cell address (None for auto-detect)
            column_only: If True, read only the first column

        Returns:
            List of numeric values, skipping blanks and non-numeric cells
        """
        if HAS_PANDAS:
            return self._read_with_pandas(sheet_name, start_cell, end_cell, column_only)
        elif HAS_OPENPYXL and self.is_xlsx:
            return self._read_with_openpyxl(sheet_name, start_cell, end_cell, column_only)
        elif HAS_XLRD and not self.is_xlsx:
            return self._read_with_xlrd(sheet_name, start_cell, end_cell, column_only)
        else:
            raise ExcelIntegrationError("No suitable Excel library available")

    def _read_with_pandas(
        self,
        sheet_name: Optional[str],
        start_cell: str,
        end_cell: Optional[str],
        column_only: bool
    ) -> List[float]:
        """Read data using pandas."""
        try:
            # Parse start cell
            start_col, start_row = A1Converter.parse_a1(start_cell)

            # Read the sheet
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                header=None,
                skiprows=start_row - 1
            )

            # Extract column data
            if column_only:
                # Read only the specified column
                col_idx = start_col - 1  # Convert to 0-based
                if col_idx >= len(df.columns):
                    return []
                series = df.iloc[:, col_idx]
            else:
                # Read entire range (not implemented for simplicity)
                col_idx = start_col - 1
                series = df.iloc[:, col_idx]

            # Convert to list of floats, filtering out invalid values
            values = []
            for val in series:
                if pd.isna(val):
                    continue
                try:
                    if isinstance(val, str):
                        # Clean string values
                        clean_val = val.strip().replace(',', '').replace('$', '').replace('%', '')
                        if clean_val == '':
                            continue
                        values.append(float(clean_val))
                    else:
                        values.append(float(val))
                except (ValueError, TypeError):
                    continue

            return values

        except Exception as e:
            raise ExcelIntegrationError(f"Pandas read error: {e}")

    def _read_with_openpyxl(
        self,
        sheet_name: Optional[str],
        start_cell: str,
        end_cell: Optional[str],
        column_only: bool
    ) -> List[float]:
        """Read data using openpyxl."""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)

            # Select worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f"Sheet '{sheet_name}' not found")
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Parse start cell
            start_col, start_row = A1Converter.parse_a1(start_cell)

            values = []
            col_letter = A1Converter.index_to_col(start_col)

            # Read column data
            for row in range(start_row, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                val = cell.value

                if val is None:
                    continue

                try:
                    if isinstance(val, str):
                        # Clean string values
                        clean_val = val.strip().replace(',', '').replace('$', '').replace('%', '')
                        if clean_val == '':
                            continue
                        values.append(float(clean_val))
                    else:
                        values.append(float(val))
                except (ValueError, TypeError):
                    continue

            wb.close()
            return values

        except Exception as e:
            raise ExcelIntegrationError(f"openpyxl read error: {e}")

    def _read_with_xlrd(
        self,
        sheet_name: Optional[str],
        start_cell: str,
        end_cell: Optional[str],
        column_only: bool
    ) -> List[float]:
        """Read data using xlrd."""
        try:
            with xlrd.open_workbook(self.file_path) as wb:
                # Select worksheet
                if sheet_name:
                    if sheet_name not in wb.sheet_names():
                        raise ValueError(f"Sheet '{sheet_name}' not found")
                    ws = wb.sheet_by_name(sheet_name)
                else:
                    ws = wb.sheet_by_index(0)

                # Parse start cell
                start_col, start_row = A1Converter.parse_a1(start_cell)
                start_col -= 1  # Convert to 0-based
                start_row -= 1

                values = []

                # Read column data
                for row in range(start_row, ws.nrows):
                    if start_col >= ws.ncols:
                        break

                    cell = ws.cell(row, start_col)
                    val = cell.value

                    if val == '' or val is None:
                        continue

                    try:
                        if cell.ctype == xlrd.XL_CELL_TEXT:
                            # Text cell - try to parse as number
                            clean_val = str(val).strip().replace(',', '').replace('$', '').replace('%', '')
                            if clean_val == '':
                                continue
                            values.append(float(clean_val))
                        elif cell.ctype in [xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_DATE]:
                            values.append(float(val))
                    except (ValueError, TypeError):
                        continue

                return values

        except Exception as e:
            raise ExcelIntegrationError(f"xlrd read error: {e}")


class ExcelWriter:
    """Write results to Excel files."""

    def __init__(self, file_path: Path, template_path: Optional[Path] = None):
        """
        Initialize Excel writer.

        Args:
            file_path: Output file path
            template_path: Optional template file to copy from
        """
        self.file_path = Path(file_path)
        self.template_path = Path(template_path) if template_path else None
        self.logger = logging.getLogger(__name__)

        if not HAS_OPENPYXL:
            raise ExcelIntegrationError("openpyxl required for Excel writing")

    def write_results(
        self,
        result: SolverResult,
        start_address: str = "A1",
        sheet_name: str = "FuzzySum Results",
        highlight_selected: bool = True,
        include_summary: bool = True
    ) -> None:
        """
        Write solver results to Excel file.

        Args:
            result: Solver result to write
            start_address: Starting address for data
            sheet_name: Sheet name for results
            highlight_selected: Whether to highlight selected cells
            include_summary: Whether to include summary table
        """
        try:
            # Create or load workbook
            if self.template_path and self.template_path.exists():
                wb = openpyxl.load_workbook(self.template_path)
            else:
                wb = openpyxl.Workbook()

            # Create or get worksheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)

            # Write summary if requested
            if include_summary:
                self._write_summary(ws, result)

            # Write selected items with highlighting
            if result.success and highlight_selected:
                self._highlight_selected_items(ws, result, start_address)

            # Save workbook
            wb.save(self.file_path)
            self.logger.info(f"Results written to {self.file_path}")

        except Exception as e:
            raise ExcelIntegrationError(f"Error writing Excel file: {e}")

    def _write_summary(self, ws: Worksheet, result: SolverResult) -> None:
        """Write summary table to worksheet."""
        # Summary table starting at A1
        headers = [
            ("Metric", "Value", "Details"),
            ("Status", "SUCCESS" if result.success else "FAILED", result.solver_status),
            ("Target Sum", f"{result.target:,.2f}", f"({result.target:.10g})"),
            ("Achieved Sum", f"{result.total_sum:,.2f}", f"({result.total_sum:.10g})"),
            ("Error", f"{result.error:,.4f}", f"{(result.error/result.target*100):.2f}% of target" if result.target != 0 else "N/A"),
            ("Items Selected", str(len(result.selected_indices)), ""),
            ("Solve Time", f"{result.solve_time:.3f}s", ""),
            ("Strategy", result.strategy_used.value, ""),
            ("Mode", result.mode_used.value, ""),
        ]

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Write headers and data
        for row_idx, row_data in enumerate(headers, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                if row_idx == 1:  # Header row
                    cell.font = header_font
                    cell.fill = header_fill
                elif row_idx == 2:  # Status row
                    if result.success:
                        cell.fill = success_fill
                    else:
                        cell.fill = error_fill

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

    def _highlight_selected_items(
        self,
        ws: Worksheet,
        result: SolverResult,
        start_address: str
    ) -> None:
        """Highlight selected items in the worksheet."""
        if not result.selected_indices:
            return

        # Parse start address
        start_col, start_row = A1Converter.parse_a1(start_address)
        col_letter = A1Converter.index_to_col(start_col)

        # Highlight style
        highlight_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        highlight_font = Font(bold=True)

        # Highlight selected cells
        for idx in result.selected_indices:
            cell_address = f"{col_letter}{start_row + idx}"
            cell = ws[cell_address]
            cell.fill = highlight_fill
            cell.font = highlight_font

            # Add comment with selection info
            if not cell.comment:
                cell.comment = openpyxl.comments.Comment(
                    f"Selected by FuzzySum\nValue: {result.selected_values[result.selected_indices.index(idx)]:.4f}",
                    "FuzzySum"
                )

    def write_batch_results(
        self,
        results: List[Tuple[float, SolverResult]],
        sheet_name: str = "Batch Results"
    ) -> None:
        """
        Write batch processing results to Excel.

        Args:
            results: List of (target, result) tuples
            sheet_name: Sheet name for batch results
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.create_sheet(sheet_name)

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Headers
            headers = [
                "Target", "Status", "Achieved Sum", "Error", "Error %",
                "Items Count", "Solve Time", "Strategy", "Mode"
            ]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")

            # Data rows
            for row_idx, (target, result) in enumerate(results, 2):
                data = [
                    target,
                    "SUCCESS" if result.success else "FAILED",
                    result.total_sum,
                    result.error,
                    (result.error / target * 100) if target != 0 else 0,
                    len(result.selected_indices),
                    result.solve_time,
                    result.strategy_used.value,
                    result.mode_used.value
                ]

                for col_idx, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value

                    # Color-code status
                    if col_idx == 2:  # Status column
                        if result.success:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Auto-size columns
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column].width = adjusted_width

            # Save workbook
            wb.save(self.file_path)
            self.logger.info(f"Batch results written to {self.file_path}")

        except Exception as e:
            raise ExcelIntegrationError(f"Error writing batch results: {e}")


def read_excel_data(
    file_path: Path,
    sheet_name: Optional[str] = None,
    start_address: str = "A1"
) -> List[float]:
    """
    Convenience function to read Excel data.

    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name (None for first sheet)
        start_address: Starting cell address

    Returns:
        List of numeric values
    """
    reader = ExcelReader(file_path)
    return reader.read_range(sheet_name, start_address)


def write_excel_results(
    result: SolverResult,
    file_path: Path,
    start_address: str = "A1",
    sheet_name: str = "FuzzySum Results"
) -> None:
    """
    Convenience function to write Excel results.

    Args:
        result: Solver result
        file_path: Output file path
        start_address: Starting address for highlighting
        sheet_name: Sheet name
    """
    writer = ExcelWriter(file_path)
    writer.write_results(result, start_address, sheet_name)


def get_available_engines() -> Dict[str, bool]:
    """Get information about available Excel engines."""
    return {
        "pandas": HAS_PANDAS,
        "openpyxl": HAS_OPENPYXL,
        "xlrd": HAS_XLRD
    }