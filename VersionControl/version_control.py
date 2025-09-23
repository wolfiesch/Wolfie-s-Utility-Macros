#!/usr/bin/env python3
"""
Version Control System for Excel Databooks
Main controller module that handles version creation, management, and coordination
"""

import argparse
import json
import shutil
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
import yaml
import sys
import os

# Add the VersionControl directory to Python path for imports
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

from metrics_extractor import MetricsExtractor
from comparator import WorkbookComparator
from storage_manager import StorageManager


class VersionController:
    """Main version control system coordinator"""

    def __init__(self, workbook_path: str, config_path: Optional[str] = None):
        """
        Initialize version controller

        Args:
            workbook_path: Path to the Excel workbook
            config_path: Optional path to config file
        """
        self.workbook_path = Path(workbook_path)
        self.project_name = self.workbook_path.stem

        # Load configuration
        if config_path:
            self.config_path = Path(config_path)
        else:
            self.config_path = Path(__file__).parent / "config.yaml"

        self.config = self._load_config()

        # Initialize components
        self.storage = StorageManager(self.project_name, self.config)
        self.setup_logging()

        # Ensure directories exist
        self._setup_directories()

    def _load_config(self) -> Dict[str, Any]:
        """Load configuration from YAML file"""
        try:
            with open(self.config_path, 'r') as f:
                return yaml.safe_load(f)
        except Exception as e:
            # Return default config if file not found
            return self._get_default_config()

    def _get_default_config(self) -> Dict[str, Any]:
        """Return default configuration"""
        return {
            "version_control": {"max_versions": 50, "auto_cleanup": True},
            "metrics": {"locations": {}},
            "comparison": {"tolerance": 0.01, "ignore_sheets": []},
            "storage": {"compression": True, "optimize_snapshots": True}
        }

    def setup_logging(self):
        """Set up logging configuration"""
        log_dir = Path("VersionControl") / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_dir / f"version_control_{datetime.now().strftime('%Y%m%d')}.log"),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def _setup_directories(self):
        """Create necessary directory structure"""
        base_dir = Path("Versions")
        self.versions_dir = base_dir / self.project_name
        self.reports_dir = Path("Reports")

        for directory in [self.versions_dir, self.reports_dir]:
            directory.mkdir(parents=True, exist_ok=True)

    def create_snapshot(self, notes: str = "", quick_save: bool = False) -> Dict[str, Any]:
        """
        Create a new version snapshot

        Args:
            notes: User notes for this version
            quick_save: If True, skip optimization for speed

        Returns:
            Dictionary with version information
        """
        try:
            self.logger.info(f"Creating snapshot for {self.project_name}")

            # Validate workbook exists
            if not self.workbook_path.exists():
                raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

            # Get next version number
            version_num = self.storage.get_next_version_number()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            version_name = f"v{version_num:03d}"

            # Create snapshot filename
            snapshot_filename = f"{version_name}_{timestamp}.xlsx"
            snapshot_path = self.versions_dir / snapshot_filename

            # Extract metrics before creating snapshot
            self.logger.info("Extracting metrics...")
            metrics_extractor = MetricsExtractor(str(self.workbook_path), self.config)
            metrics = metrics_extractor.extract()

            # Create optimized snapshot
            if quick_save:
                # Quick copy without optimization
                shutil.copy2(self.workbook_path, snapshot_path)
            else:
                # Optimized copy with cleanup
                self._create_optimized_snapshot(snapshot_path)

            # Calculate file hash
            file_hash = self._calculate_file_hash(snapshot_path)

            # Create version metadata
            version_info = {
                "version": version_name,
                "timestamp": timestamp,
                "datetime": datetime.now().isoformat(),
                "file_size": snapshot_path.stat().st_size,
                "file_path": str(snapshot_path),
                "metrics": metrics,
                "notes": notes,
                "hash": file_hash,
                "quick_save": quick_save
            }

            # Save metadata
            self.storage.save_version_metadata(version_info)

            # Cleanup old versions if needed
            if self.config["version_control"]["auto_cleanup"]:
                self._cleanup_old_versions()

            self.logger.info(f"Snapshot created successfully: {version_name}")

            return {
                "success": True,
                "version": version_name,
                "path": str(snapshot_path),
                "file_size": version_info["file_size"],
                "metrics": metrics
            }

        except Exception as e:
            self.logger.error(f"Failed to create snapshot: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def _create_optimized_snapshot(self, snapshot_path: Path):
        """Create optimized snapshot by removing unused areas"""
        from openpyxl import load_workbook

        # Load workbook
        wb = load_workbook(str(self.workbook_path))

        if self.config["storage"]["optimize_snapshots"]:
            # Remove empty rows and columns from each sheet
            for sheet in wb.worksheets:
                # Get used range
                if sheet.max_row > 1 and sheet.max_column > 1:
                    # Remove empty rows at the end
                    for row in range(sheet.max_row, 0, -1):
                        if any(sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)):
                            break
                        sheet.delete_rows(row)

                    # Remove empty columns at the end
                    for col in range(sheet.max_column, 0, -1):
                        if any(sheet.cell(row, col).value for row in range(1, sheet.max_row + 1)):
                            break
                        sheet.delete_cols(col)

        # Save optimized workbook
        wb.save(str(snapshot_path))
        wb.close()

    def _calculate_file_hash(self, file_path: Path) -> str:
        """Calculate SHA-256 hash of file"""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def list_versions(self) -> List[Dict[str, Any]]:
        """Get list of all versions for this project"""
        try:
            versions = self.storage.get_all_versions()

            # Format for Excel dropdown
            version_list = []
            for version in versions:
                display_text = f"{version['version']} - {version['timestamp']}"
                if version.get('notes'):
                    display_text += f" - {version['notes'][:30]}"
                    if len(version['notes']) > 30:
                        display_text += "..."

                version_list.append({
                    "value": version['version'],
                    "display": display_text,
                    "timestamp": version['timestamp'],
                    "file_size": version['file_size'],
                    "metrics": version.get('metrics', {})
                })

            return version_list

        except Exception as e:
            self.logger.error(f"Failed to list versions: {str(e)}")
            return []

    def compare_to_version(self, version_name: str) -> Dict[str, Any]:
        """Compare current workbook to a specific version"""
        try:
            self.logger.info(f"Comparing current workbook to {version_name}")

            # Get version file path
            version_info = self.storage.get_version_info(version_name)
            if not version_info:
                raise ValueError(f"Version {version_name} not found")

            version_path = Path(version_info['file_path'])
            if not version_path.exists():
                raise FileNotFoundError(f"Version file not found: {version_path}")

            # Create comparator
            comparator = WorkbookComparator(
                str(self.workbook_path),
                str(version_path),
                self.config
            )

            # Perform comparison
            comparison_results = comparator.compare()

            # Generate comparison report
            report_path = self._generate_comparison_report(comparison_results, version_name)

            return {
                "success": True,
                "comparison_results": comparison_results,
                "report_path": str(report_path),
                "version_compared": version_name
            }

        except Exception as e:
            self.logger.error(f"Failed to compare to version {version_name}: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def _generate_comparison_report(self, results: Dict[str, Any], version_name: str) -> Path:
        """Generate Excel comparison report"""
        import pandas as pd
        from openpyxl.styles import PatternFill

        # Create report filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_filename = f"comparison_{self.project_name}_{version_name}_{timestamp}.xlsx"
        report_path = self.reports_dir / report_filename

        with pd.ExcelWriter(str(report_path), engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                "Metric": ["Total Changes", "Sheets Modified", "Major Changes", "Formula Changes"],
                "Count": [
                    len(results.get('cell_changes', [])),
                    len(set(change['sheet'] for change in results.get('cell_changes', []))),
                    len([c for c in results.get('cell_changes', []) if c.get('major_change', False)]),
                    len([c for c in results.get('cell_changes', []) if c.get('formula_change', False)])
                ]
            }

            if 'metrics_comparison' in results:
                summary_data["Metric"].extend(results['metrics_comparison'].keys())
                summary_data["Count"].extend([
                    f"{mc.get('old', 'N/A')} → {mc.get('new', 'N/A')}"
                    for mc in results['metrics_comparison'].values()
                ])

            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Cell changes detail
            if results.get('cell_changes'):
                changes_df = pd.DataFrame(results['cell_changes'])
                changes_df.to_excel(writer, sheet_name='Cell Changes', index=False)

                # Apply formatting
                worksheet = writer.sheets['Cell Changes']
                for idx, change in enumerate(results['cell_changes'], 2):  # Start from row 2 (after header)
                    if change.get('major_change', False):
                        fill = PatternFill(start_color='FFAAAA', end_color='FFAAAA', fill_type='solid')
                        for col in range(1, len(changes_df.columns) + 1):
                            worksheet.cell(row=idx, column=col).fill = fill

        return report_path

    def rollback_to_version(self, version_name: str, backup_current: bool = True) -> Dict[str, Any]:
        """Rollback to a specific version"""
        try:
            self.logger.info(f"Rolling back to version {version_name}")

            # Backup current version if requested
            if backup_current:
                backup_result = self.create_snapshot(f"Backup before rollback to {version_name}", quick_save=True)
                if not backup_result['success']:
                    raise Exception(f"Failed to create backup: {backup_result['error']}")

            # Get version info
            version_info = self.storage.get_version_info(version_name)
            if not version_info:
                raise ValueError(f"Version {version_name} not found")

            version_path = Path(version_info['file_path'])
            if not version_path.exists():
                raise FileNotFoundError(f"Version file not found: {version_path}")

            # Close current workbook (this would need to be handled by VBA)
            # For now, just copy the file
            shutil.copy2(version_path, self.workbook_path)

            self.logger.info(f"Successfully rolled back to {version_name}")

            return {
                "success": True,
                "rolled_back_to": version_name,
                "backup_created": backup_current
            }

        except Exception as e:
            self.logger.error(f"Failed to rollback to {version_name}: {str(e)}")
            return {
                "success": False,
                "error": str(e)
            }

    def _cleanup_old_versions(self):
        """Remove old versions beyond the maximum limit"""
        max_versions = self.config["version_control"]["max_versions"]
        versions = self.storage.get_all_versions()

        if len(versions) > max_versions:
            # Sort by timestamp (oldest first)
            versions.sort(key=lambda x: x['timestamp'])

            # Remove oldest versions
            versions_to_remove = versions[:-max_versions]
            for version in versions_to_remove:
                try:
                    version_path = Path(version['file_path'])
                    if version_path.exists():
                        version_path.unlink()
                    self.storage.remove_version_metadata(version['version'])
                    self.logger.info(f"Cleaned up old version: {version['version']}")
                except Exception as e:
                    self.logger.warning(f"Failed to cleanup version {version['version']}: {str(e)}")

    def get_project_stats(self) -> Dict[str, Any]:
        """Get statistics about the project"""
        versions = self.storage.get_all_versions()

        if not versions:
            return {"total_versions": 0}

        # Calculate statistics
        total_size = sum(v['file_size'] for v in versions)
        latest_version = max(versions, key=lambda x: x['timestamp'])

        return {
            "total_versions": len(versions),
            "total_size_mb": round(total_size / (1024 * 1024), 2),
            "latest_version": latest_version['version'],
            "latest_timestamp": latest_version['timestamp'],
            "project_name": self.project_name
        }


def main():
    """Command line interface for the version control system"""
    parser = argparse.ArgumentParser(description='Excel Databook Version Control System')
    parser.add_argument('--action', required=True,
                       choices=['create_snapshot', 'list_versions', 'compare', 'rollback', 'stats'],
                       help='Action to perform')
    parser.add_argument('--workbook', required=True, help='Path to Excel workbook')
    parser.add_argument('--version', help='Version name for comparison or rollback')
    parser.add_argument('--notes', default='', help='Notes for snapshot creation')
    parser.add_argument('--quick', action='store_true', help='Quick save (skip optimization)')
    parser.add_argument('--config', help='Path to config file')

    args = parser.parse_args()

    # Initialize version controller
    vc = VersionController(args.workbook, args.config)

    # Execute requested action
    try:
        if args.action == 'create_snapshot':
            result = vc.create_snapshot(args.notes, args.quick)

        elif args.action == 'list_versions':
            result = vc.list_versions()

        elif args.action == 'compare':
            if not args.version:
                raise ValueError("Version name required for comparison")
            result = vc.compare_to_version(args.version)

        elif args.action == 'rollback':
            if not args.version:
                raise ValueError("Version name required for rollback")
            result = vc.rollback_to_version(args.version)

        elif args.action == 'stats':
            result = vc.get_project_stats()

        # Output result as JSON for VBA consumption
        print(json.dumps(result, indent=2))

    except Exception as e:
        error_result = {"success": False, "error": str(e)}
        print(json.dumps(error_result, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()