#!/usr/bin/env python3
"""
Metrics Extraction Engine for Excel Databooks
Intelligently extracts key financial metrics from standardized databook layouts
"""

import re
import logging
from typing import Dict, List, Optional, Any, Tuple
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


class MetricsExtractor:
    """Intelligent extraction of key financial metrics from Excel databooks"""

    def __init__(self, workbook_path: str, config: Dict[str, Any]):
        """
        Initialize metrics extractor

        Args:
            workbook_path: Path to Excel workbook
            config: Configuration dictionary with metric locations
        """
        self.workbook_path = Path(workbook_path)
        self.config = config
        self.logger = logging.getLogger(__name__)

        # Load workbook
        try:
            self.workbook = openpyxl.load_workbook(str(workbook_path), data_only=True)
        except Exception as e:
            self.logger.error(f"Failed to load workbook {workbook_path}: {str(e)}")
            raise

        # Get metrics configuration
        self.metrics_config = config.get("metrics", {}).get("locations", {})

    def extract(self) -> Dict[str, Any]:
        """
        Extract all configured metrics from the workbook

        Returns:
            Dictionary containing extracted metrics and metadata
        """
        extracted_metrics = {}

        try:
            # Extract each configured metric
            for metric_name, metric_config in self.metrics_config.items():
                self.logger.info(f"Extracting metric: {metric_name}")
                value = self._extract_metric(metric_name, metric_config)
                extracted_metrics[metric_name] = value

            # Add databook statistics
            extracted_metrics.update(self._extract_databook_stats())

            # Add extraction metadata
            extracted_metrics["_metadata"] = {
                "extraction_timestamp": self._get_timestamp(),
                "workbook_name": self.workbook_path.name,
                "sheets_analyzed": len(self.workbook.worksheets),
                "extraction_success": True
            }

        except Exception as e:
            self.logger.error(f"Error during metrics extraction: {str(e)}")
            extracted_metrics["_metadata"] = {
                "extraction_timestamp": self._get_timestamp(),
                "extraction_success": False,
                "error": str(e)
            }

        finally:
            self.workbook.close()

        return extracted_metrics

    def _extract_metric(self, metric_name: str, config: Dict[str, Any]) -> Optional[float]:
        """
        Extract a specific metric using its configuration

        Args:
            metric_name: Name of the metric to extract
            config: Configuration for this metric

        Returns:
            Extracted value or None if not found
        """
        sheets_to_search = config.get("sheets", [])
        patterns = config.get("patterns", [])
        search_range = config.get("search_range", "A1:Z100")

        for sheet_name in sheets_to_search:
            if sheet_name in self.workbook.sheetnames:
                sheet = self.workbook[sheet_name]
                value = self._search_sheet_for_metric(sheet, patterns, search_range)
                if value is not None:
                    self.logger.info(f"Found {metric_name} in sheet {sheet_name}: {value}")
                    return value

        # If not found in configured sheets, try intelligent search
        return self._intelligent_metric_search(metric_name, patterns)

    def _search_sheet_for_metric(self, sheet, patterns: List[str], search_range: str) -> Optional[float]:
        """
        Search for metric in a specific sheet

        Args:
            sheet: Worksheet to search
            patterns: List of text patterns to look for
            search_range: Excel range to search (e.g., "A1:Z100")

        Returns:
            Found value or None
        """
        try:
            # Parse search range
            start_cell, end_cell = search_range.split(":")
            start_col, start_row = self._parse_cell_reference(start_cell)
            end_col, end_row = self._parse_cell_reference(end_cell)

            # Search for patterns
            for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
                for col in range(start_col, min(end_col + 1, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)

                    if cell.value and isinstance(cell.value, str):
                        cell_text = cell.value.upper().strip()

                        # Check if any pattern matches
                        for pattern in patterns:
                            if pattern.upper() in cell_text:
                                # Found pattern, look for value nearby
                                value = self._find_value_near_cell(sheet, row, col)
                                if value is not None:
                                    return value

        except Exception as e:
            self.logger.warning(f"Error searching sheet {sheet.title}: {str(e)}")

        return None

    def _find_value_near_cell(self, sheet, label_row: int, label_col: int) -> Optional[float]:
        """
        Find numeric value near a label cell

        Args:
            sheet: Worksheet
            label_row: Row of the label cell
            label_col: Column of the label cell

        Returns:
            Found numeric value or None
        """
        # Search pattern: right, down, down-right, right-right
        search_offsets = [
            (0, 1), (0, 2), (0, 3), (0, 4), (0, 5),  # Right
            (1, 0), (2, 0), (3, 0), (4, 0),           # Down
            (1, 1), (1, 2), (2, 1),                   # Down-right
            (-1, 1), (-1, 2)                          # Up-right
        ]

        for row_offset, col_offset in search_offsets:
            try:
                search_row = label_row + row_offset
                search_col = label_col + col_offset

                if search_row > 0 and search_col > 0:
                    cell = sheet.cell(row=search_row, column=search_col)
                    value = self._extract_numeric_value(cell.value)
                    if value is not None:
                        return value

            except Exception:
                continue

        return None

    def _extract_numeric_value(self, cell_value) -> Optional[float]:
        """
        Extract numeric value from cell content

        Args:
            cell_value: Cell value to process

        Returns:
            Numeric value or None
        """
        if cell_value is None:
            return None

        # If already a number
        if isinstance(cell_value, (int, float)):
            return float(cell_value)

        # If string, try to extract number
        if isinstance(cell_value, str):
            # Remove common formatting characters
            cleaned = re.sub(r'[,$%\s()]', '', cell_value)

            # Handle parentheses for negative numbers
            if cell_value.strip().startswith('(') and cell_value.strip().endswith(')'):
                cleaned = '-' + cleaned

            try:
                return float(cleaned)
            except ValueError:
                pass

        return None

    def _parse_cell_reference(self, cell_ref: str) -> Tuple[int, int]:
        """
        Parse Excel cell reference into column and row numbers

        Args:
            cell_ref: Cell reference like "A1" or "AB123"

        Returns:
            Tuple of (column_number, row_number)
        """
        match = re.match(r'([A-Z]+)(\d+)', cell_ref.upper())
        if match:
            col_letters = match.group(1)
            row_num = int(match.group(2))
            col_num = column_index_from_string(col_letters)
            return col_num, row_num
        else:
            raise ValueError(f"Invalid cell reference: {cell_ref}")

    def _intelligent_metric_search(self, metric_name: str, patterns: List[str]) -> Optional[float]:
        """
        Perform intelligent search across all sheets when metric not found in configured locations

        Args:
            metric_name: Name of metric to search for
            patterns: Text patterns to search for

        Returns:
            Found value or None
        """
        self.logger.info(f"Performing intelligent search for {metric_name}")

        # Define sheet priority for different metric types
        sheet_priorities = {
            "ebitda": ["EBITDA", "P&L", "Income", "Profit"],
            "revenue": ["P&L", "Income", "Profit", "Detail_PL"],
            "total_assets": ["Balance", "BS", "Detail_BS"],
            "total_liabilities": ["Balance", "BS", "Detail_BS"],
            "working_capital": ["NWC", "Working", "Balance"],
            "net_debt": ["Debt", "NWC", "Balance"]
        }

        # Get relevant sheet names
        relevant_keywords = sheet_priorities.get(metric_name, [])
        candidate_sheets = []

        for sheet_name in self.workbook.sheetnames:
            # Skip summary/index sheets
            if any(skip in sheet_name.upper() for skip in ["INDEX", "COVER", "SUMMARY", "TOC"]):
                continue

            # Prioritize sheets with relevant keywords
            priority = 0
            for keyword in relevant_keywords:
                if keyword.upper() in sheet_name.upper():
                    priority += 10

            # Add detail sheets with higher priority
            if "DETAIL" in sheet_name.upper():
                priority += 5

            if priority > 0 or not candidate_sheets:  # Include all sheets if no priorities match
                candidate_sheets.append((priority, sheet_name))

        # Sort by priority (highest first)
        candidate_sheets.sort(reverse=True, key=lambda x: x[0])

        # Search candidate sheets
        for priority, sheet_name in candidate_sheets:
            sheet = self.workbook[sheet_name]
            value = self._search_sheet_for_metric(sheet, patterns, "A1:AZ200")  # Broader search
            if value is not None:
                self.logger.info(f"Intelligent search found {metric_name} in {sheet_name}: {value}")
                return value

        self.logger.warning(f"Could not find metric {metric_name} with patterns {patterns}")
        return None

    def _extract_databook_stats(self) -> Dict[str, Any]:
        """
        Extract general databook statistics

        Returns:
            Dictionary with databook statistics
        """
        stats = {}

        try:
            # Sheet count
            stats["sheet_count"] = len(self.workbook.worksheets)

            # Formula count
            stats["formula_count"] = self._count_formulas()

            # Identify sheet types
            sheet_types = self._classify_sheets()
            stats["sheet_types"] = sheet_types

            # File size (approximate)
            stats["file_size_bytes"] = self.workbook_path.stat().st_size

            # Last modified
            stats["last_modified"] = self.workbook_path.stat().st_mtime

        except Exception as e:
            self.logger.warning(f"Error extracting databook stats: {str(e)}")

        return stats

    def _count_formulas(self) -> int:
        """Count total number of formulas in the workbook"""
        formula_count = 0

        try:
            # Reload workbook with formulas (not values)
            formula_wb = openpyxl.load_workbook(str(self.workbook_path), data_only=False)

            for sheet in formula_wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula cell
                            formula_count += 1

            formula_wb.close()

        except Exception as e:
            self.logger.warning(f"Error counting formulas: {str(e)}")

        return formula_count

    def _classify_sheets(self) -> Dict[str, List[str]]:
        """
        Classify sheets by type (EBITDA, P&L, Balance Sheet, etc.)

        Returns:
            Dictionary with sheet classifications
        """
        classifications = {
            "ebitda": [],
            "profit_loss": [],
            "balance_sheet": [],
            "cash_flow": [],
            "working_capital": [],
            "detail": [],
            "summary": [],
            "database": [],
            "other": []
        }

        for sheet_name in self.workbook.sheetnames:
            name_upper = sheet_name.upper()

            # Classify by keywords
            if any(kw in name_upper for kw in ["EBITDA"]):
                classifications["ebitda"].append(sheet_name)
            elif any(kw in name_upper for kw in ["P&L", "PL", "PROFIT", "INCOME"]):
                classifications["profit_loss"].append(sheet_name)
            elif any(kw in name_upper for kw in ["BS", "BALANCE"]):
                classifications["balance_sheet"].append(sheet_name)
            elif any(kw in name_upper for kw in ["CASH", "CF"]):
                classifications["cash_flow"].append(sheet_name)
            elif any(kw in name_upper for kw in ["NWC", "WORKING"]):
                classifications["working_capital"].append(sheet_name)
            elif "DETAIL" in name_upper:
                classifications["detail"].append(sheet_name)
            elif "SUMMARY" in name_upper:
                classifications["summary"].append(sheet_name)
            elif "_DB" in name_upper or "DATABASE" in name_upper:
                classifications["database"].append(sheet_name)
            else:
                classifications["other"].append(sheet_name)

        return classifications

    def _get_timestamp(self) -> str:
        """Get current timestamp as ISO string"""
        from datetime import datetime
        return datetime.now().isoformat()

    def extract_custom_metric(self, sheet_name: str, cell_reference: str) -> Optional[float]:
        """
        Extract a custom metric from a specific location

        Args:
            sheet_name: Name of the sheet
            cell_reference: Cell reference (e.g., "B15")

        Returns:
            Extracted value or None
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                return None

            sheet = self.workbook[sheet_name]
            cell = sheet[cell_reference]
            return self._extract_numeric_value(cell.value)

        except Exception as e:
            self.logger.warning(f"Error extracting custom metric from {sheet_name}!{cell_reference}: {str(e)}")
            return None

    def search_for_pattern(self, pattern: str, sheet_names: Optional[List[str]] = None) -> List[Dict[str, Any]]:
        """
        Search for a specific pattern across sheets

        Args:
            pattern: Text pattern to search for
            sheet_names: Optional list of sheets to search (None for all)

        Returns:
            List of dictionaries with found locations and values
        """
        results = []
        sheets_to_search = sheet_names or self.workbook.sheetnames

        for sheet_name in sheets_to_search:
            if sheet_name not in self.workbook.sheetnames:
                continue

            sheet = self.workbook[sheet_name]

            for row in range(1, min(sheet.max_row + 1, 200)):  # Limit search for performance
                for col in range(1, min(sheet.max_column + 1, 50)):
                    cell = sheet.cell(row=row, column=col)

                    if cell.value and isinstance(cell.value, str):
                        if pattern.upper() in cell.value.upper():
                            # Found pattern, get nearby value
                            value = self._find_value_near_cell(sheet, row, col)
                            if value is not None:
                                results.append({
                                    "sheet": sheet_name,
                                    "cell": f"{get_column_letter(col)}{row}",
                                    "label": cell.value,
                                    "value": value,
                                    "row": row,
                                    "column": col
                                })

        return results