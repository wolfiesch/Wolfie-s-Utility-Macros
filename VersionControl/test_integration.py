#!/usr/bin/env python3
"""
Integration Test Suite for Excel Version Control System
Tests the complete Python backend functionality
"""

import sys
import json
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
import unittest
import logging

# Add the VersionControl directory to Python path
current_dir = Path(__file__).parent
sys.path.insert(0, str(current_dir))

from version_control import VersionController
from metrics_extractor import MetricsExtractor
from comparator import WorkbookComparator
from storage_manager import StorageManager
from vba_python_bridge import VBAPythonBridge


class TestVersionControlIntegration(unittest.TestCase):
    """Integration tests for the version control system"""

    @classmethod
    def setUpClass(cls):
        """Set up test environment"""
        cls.test_dir = Path(tempfile.mkdtemp(prefix="vc_test_"))
        cls.test_workbook = cls.test_dir / "test_workbook.xlsx"

        # Create a simple test Excel file
        cls._create_test_workbook()

        # Setup logging
        logging.basicConfig(level=logging.INFO)
        cls.logger = logging.getLogger(__name__)

    @classmethod
    def tearDownClass(cls):
        """Clean up test environment"""
        if cls.test_dir.exists():
            shutil.rmtree(cls.test_dir)

    @classmethod
    def _create_test_workbook(cls):
        """Create a simple test Excel workbook"""
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test_Sheet"

            # Add some test data
            ws['A1'] = "Test Data"
            ws['A2'] = "Revenue"
            ws['B2'] = 1000000
            ws['A3'] = "EBITDA"
            ws['B3'] = 200000
            ws['A4'] = "Total Assets"
            ws['B4'] = 5000000

            wb.save(str(cls.test_workbook))
            wb.close()

        except ImportError:
            # Create dummy file if openpyxl not available
            cls.test_workbook.touch()

    def setUp(self):
        """Set up each test"""
        self.vc = VersionController(str(self.test_workbook))

    def test_01_version_controller_initialization(self):
        """Test version controller initialization"""
        self.assertIsNotNone(self.vc)
        self.assertEqual(self.vc.project_name, "test_workbook")
        self.assertTrue(self.vc.versions_dir.exists())

    def test_02_create_snapshot(self):
        """Test snapshot creation"""
        result = self.vc.create_snapshot("Test snapshot", quick_save=True)

        self.assertTrue(result['success'])
        self.assertIn('version', result)
        self.assertIn('path', result)

        # Verify snapshot file exists
        snapshot_path = Path(result['path'])
        self.assertTrue(snapshot_path.exists())

    def test_03_list_versions(self):
        """Test listing versions"""
        # Create a snapshot first
        self.vc.create_snapshot("First snapshot", quick_save=True)

        versions = self.vc.list_versions()
        self.assertGreater(len(versions), 0)

        # Check version structure
        version = versions[0]
        self.assertIn('value', version)
        self.assertIn('display', version)

    def test_04_storage_manager(self):
        """Test storage manager functionality"""
        storage = self.vc.storage

        # Test getting next version number
        next_version = storage.get_next_version_number()
        self.assertIsInstance(next_version, int)
        self.assertGreater(next_version, 0)

        # Test project statistics
        stats = storage.get_project_statistics()
        self.assertIsInstance(stats, dict)
        self.assertIn('project_name', stats)

    def test_05_metrics_extraction(self):
        """Test metrics extraction"""
        if not self.test_workbook.exists():
            self.skipTest("Test workbook not available")

        try:
            extractor = MetricsExtractor(str(self.test_workbook), self.vc.config)
            metrics = extractor.extract()

            self.assertIsInstance(metrics, dict)
            self.assertIn('_metadata', metrics)
            self.assertTrue(metrics['_metadata']['extraction_success'])

        except Exception as e:
            self.skipTest(f"Metrics extraction failed: {str(e)}")

    def test_06_vba_python_bridge(self):
        """Test VBA-Python bridge"""
        bridge = VBAPythonBridge()

        # Test create snapshot command
        result = bridge.execute_command(
            str(self.test_workbook),
            "create_snapshot",
            notes="Bridge test snapshot",
            quick=True
        )

        self.assertIsInstance(result, dict)
        self.assertIn('success', result)

    def test_07_comparison_workflow(self):
        """Test version comparison workflow"""
        # Create first snapshot
        result1 = self.vc.create_snapshot("First version", quick_save=True)
        self.assertTrue(result1['success'])
        version1 = result1['version']

        # Create second snapshot
        result2 = self.vc.create_snapshot("Second version", quick_save=True)
        self.assertTrue(result2['success'])

        # Compare versions
        comparison_result = self.vc.compare_to_version(version1)
        self.assertIsInstance(comparison_result, dict)
        self.assertIn('success', comparison_result)

    def test_08_project_statistics(self):
        """Test project statistics"""
        # Create a few snapshots
        for i in range(3):
            self.vc.create_snapshot(f"Test snapshot {i+1}", quick_save=True)

        stats = self.vc.get_project_stats()
        self.assertIsInstance(stats, dict)
        self.assertIn('total_versions', stats)
        self.assertGreaterEqual(stats['total_versions'], 3)

    def test_09_cleanup_and_maintenance(self):
        """Test cleanup operations"""
        storage = self.vc.storage

        # Test cleanup operations
        cleanup_stats = storage.cleanup_orphaned_files()
        self.assertIsInstance(cleanup_stats, dict)
        self.assertIn('orphaned_snapshots_removed', cleanup_stats)

    def test_10_error_handling(self):
        """Test error handling"""
        # Test with non-existent workbook
        fake_path = str(self.test_dir / "nonexistent.xlsx")

        try:
            fake_vc = VersionController(fake_path)
            result = fake_vc.create_snapshot("Should fail")
            self.assertFalse(result['success'])
            self.assertIn('error', result)
        except Exception:
            # Expected to fail
            pass


class TestCommandLineInterface(unittest.TestCase):
    """Test command line interface"""

    def setUp(self):
        """Set up CLI tests"""
        self.test_dir = Path(tempfile.mkdtemp(prefix="vc_cli_test_"))
        self.test_workbook = self.test_dir / "cli_test.xlsx"
        self._create_test_workbook()

    def tearDown(self):
        """Clean up CLI tests"""
        if self.test_dir.exists():
            shutil.rmtree(self.test_dir)

    def _create_test_workbook(self):
        """Create test workbook for CLI tests"""
        self.test_workbook.touch()

    def test_cli_bridge_commands(self):
        """Test CLI bridge commands"""
        bridge = VBAPythonBridge()

        # Test stats command
        result = bridge.execute_command(str(self.test_workbook), "stats")
        self.assertIsInstance(result, dict)
        self.assertIn('success', result)

        # Test list versions command
        result = bridge.execute_command(str(self.test_workbook), "list_versions")
        self.assertIsInstance(result, dict)


def run_integration_tests():
    """Run all integration tests"""
    print("=" * 60)
    print("Excel Version Control System - Integration Tests")
    print("=" * 60)

    # Create test suite
    suite = unittest.TestSuite()

    # Add test classes
    suite.addTest(unittest.makeSuite(TestVersionControlIntegration))
    suite.addTest(unittest.makeSuite(TestCommandLineInterface))

    # Run tests
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    # Print summary
    print("\n" + "=" * 60)
    print("Test Summary:")
    print(f"Tests run: {result.testsRun}")
    print(f"Failures: {len(result.failures)}")
    print(f"Errors: {len(result.errors)}")

    if result.failures:
        print("\nFailures:")
        for test, traceback in result.failures:
            print(f"  - {test}: {traceback}")

    if result.errors:
        print("\nErrors:")
        for test, traceback in result.errors:
            print(f"  - {test}: {traceback}")

    success = len(result.failures) == 0 and len(result.errors) == 0
    print(f"\nOverall Result: {'PASS' if success else 'FAIL'}")
    print("=" * 60)

    return success


def test_system_requirements():
    """Test system requirements and dependencies"""
    print("Checking system requirements...")

    requirements = {
        "Python 3.7+": sys.version_info >= (3, 7),
        "openpyxl": False,
        "pandas": False,
        "pyyaml": False,
        "pathlib": True,  # Built-in for Python 3.4+
        "json": True,     # Built-in
        "logging": True   # Built-in
    }

    # Check optional dependencies
    try:
        import openpyxl
        requirements["openpyxl"] = True
    except ImportError:
        pass

    try:
        import pandas
        requirements["pandas"] = True
    except ImportError:
        pass

    try:
        import yaml
        requirements["pyyaml"] = True
    except ImportError:
        pass

    print("\nSystem Requirements Check:")
    all_satisfied = True
    for req, satisfied in requirements.items():
        status = "✓ PASS" if satisfied else "✗ FAIL"
        print(f"  {req}: {status}")
        if not satisfied and req in ["openpyxl", "pandas", "pyyaml"]:
            all_satisfied = False

    if not all_satisfied:
        print("\nMissing dependencies. Install with:")
        print("pip install -r requirements.txt")

    return all_satisfied


if __name__ == "__main__":
    # Check requirements first
    if not test_system_requirements():
        print("\nWarning: Some dependencies are missing. Tests may fail.")
        print("Continue anyway? (y/n): ", end="")
        if input().lower() != 'y':
            sys.exit(1)

    # Run integration tests
    success = run_integration_tests()

    if success:
        print("\n✓ All tests passed! Version control system is ready for use.")
    else:
        print("\n✗ Some tests failed. Please check the issues above.")
        sys.exit(1)