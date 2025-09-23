#!/usr/bin/env python3
"""
Workbook Comparison Engine
Performs intelligent comparison between Excel workbooks with detailed change detection
"""

import logging
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Tuple, Optional
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font
import re


class WorkbookComparator:
    """Intelligent comparison engine for Excel workbooks"""

    def __init__(self, current_path: str, comparison_path: str, config: Dict[str, Any]):
        """
        Initialize workbook comparator

        Args:
            current_path: Path to current workbook
            comparison_path: Path to comparison workbook
            config: Configuration dictionary
        """
        self.current_path = Path(current_path)
        self.comparison_path = Path(comparison_path)
        self.config = config
        self.logger = logging.getLogger(__name__)

        # Comparison configuration
        self.comparison_config = config.get("comparison", {})
        self.tolerance = self.comparison_config.get("tolerance", 0.01)
        self.ignore_sheets = self.comparison_config.get("ignore_sheets", [])
        self.max_cells_per_sheet = self.comparison_config.get("max_cells_per_sheet", 10000)

        # Load workbooks
        self._load_workbooks()

    def _load_workbooks(self):
        """Load both workbooks for comparison"""
        try:
            self.current_wb = openpyxl.load_workbook(str(self.current_path), data_only=True)
            self.comparison_wb = openpyxl.load_workbook(str(self.comparison_path), data_only=True)
            self.logger.info(f"Loaded workbooks for comparison: {self.current_path.name} vs {self.comparison_path.name}")
        except Exception as e:
            self.logger.error(f"Failed to load workbooks: {str(e)}")
            raise

    def compare(self) -> Dict[str, Any]:
        """
        Perform comprehensive workbook comparison

        Returns:
            Dictionary containing comparison results
        """
        try:
            self.logger.info("Starting workbook comparison")

            comparison_results = {
                "structural_changes": self._compare_structure(),
                "cell_changes": self._compare_cell_values(),
                "formula_changes": self._compare_formulas(),
                "metrics_comparison": self._compare_metrics(),
                "summary": {}
            }

            # Generate summary
            comparison_results["summary"] = self._generate_summary(comparison_results)

            # Identify significant changes
            comparison_results["alerts"] = self._identify_alerts(comparison_results)

            self.logger.info("Workbook comparison completed")
            return comparison_results

        except Exception as e:
            self.logger.error(f"Error during workbook comparison: {str(e)}")
            return {
                "error": str(e),
                "structural_changes": {},
                "cell_changes": [],
                "formula_changes": [],
                "metrics_comparison": {},
                "summary": {"comparison_failed": True}
            }

        finally:
            self._close_workbooks()

    def _compare_structure(self) -> Dict[str, Any]:
        """Compare workbook structure (sheets, names, etc.)"""
        structural_changes = {
            "sheets_added": [],
            "sheets_removed": [],
            "sheets_renamed": [],
            "sheet_count_change": 0
        }

        try:
            current_sheets = set(self.current_wb.sheetnames)
            comparison_sheets = set(self.comparison_wb.sheetnames)

            # Find added and removed sheets
            structural_changes["sheets_added"] = list(current_sheets - comparison_sheets)
            structural_changes["sheets_removed"] = list(comparison_sheets - current_sheets)
            structural_changes["sheet_count_change"] = len(current_sheets) - len(comparison_sheets)

            # Detect potential sheet renames (similar names)
            if structural_changes["sheets_added"] and structural_changes["sheets_removed"]:
                renames = self._detect_sheet_renames(
                    structural_changes["sheets_added"],
                    structural_changes["sheets_removed"]
                )
                structural_changes["sheets_renamed"] = renames

        except Exception as e:
            self.logger.warning(f"Error comparing structure: {str(e)}")

        return structural_changes

    def _detect_sheet_renames(self, added_sheets: List[str], removed_sheets: List[str]) -> List[Dict[str, str]]:
        """Detect potential sheet renames based on similarity"""
        renames = []

        for added_sheet in added_sheets:
            best_match = None
            best_similarity = 0

            for removed_sheet in removed_sheets:
                similarity = self._calculate_string_similarity(added_sheet, removed_sheet)
                if similarity > best_similarity and similarity > 0.7:  # 70% similarity threshold
                    best_similarity = similarity
                    best_match = removed_sheet

            if best_match:
                renames.append({
                    "old_name": best_match,
                    "new_name": added_sheet,
                    "similarity": best_similarity
                })

        return renames

    def _calculate_string_similarity(self, str1: str, str2: str) -> float:
        """Calculate similarity between two strings using simple ratio"""
        if not str1 or not str2:
            return 0.0

        # Simple similarity based on character overlap
        str1_lower = str1.lower()
        str2_lower = str2.lower()

        if str1_lower == str2_lower:
            return 1.0

        # Calculate character overlap
        overlap = len(set(str1_lower) & set(str2_lower))
        total_chars = len(set(str1_lower) | set(str2_lower))

        return overlap / total_chars if total_chars > 0 else 0.0

    def _compare_cell_values(self) -> List[Dict[str, Any]]:
        """Compare cell values between workbooks"""
        cell_changes = []

        # Get common sheets (excluding ignored sheets)
        common_sheets = set(self.current_wb.sheetnames) & set(self.comparison_wb.sheetnames)
        common_sheets = [s for s in common_sheets if s not in self.ignore_sheets]

        for sheet_name in common_sheets:
            try:
                sheet_changes = self._compare_sheet_values(sheet_name)
                cell_changes.extend(sheet_changes)
            except Exception as e:
                self.logger.warning(f"Error comparing sheet {sheet_name}: {str(e)}")

        return cell_changes

    def _compare_sheet_values(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Compare values in a specific sheet"""
        changes = []

        try:
            current_sheet = self.current_wb[sheet_name]
            comparison_sheet = self.comparison_wb[sheet_name]

            # Get the overlapping range to compare
            max_row = min(current_sheet.max_row, comparison_sheet.max_row, 500)  # Limit for performance
            max_col = min(current_sheet.max_column, comparison_sheet.max_column, 100)

            cells_compared = 0

            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    if cells_compared >= self.max_cells_per_sheet:
                        break

                    try:
                        current_cell = current_sheet.cell(row=row, column=col)
                        comparison_cell = comparison_sheet.cell(row=row, column=col)

                        change_info = self._compare_cell_pair(
                            current_cell, comparison_cell, sheet_name, row, col
                        )

                        if change_info:
                            changes.append(change_info)

                        cells_compared += 1

                    except Exception as e:
                        continue

                if cells_compared >= self.max_cells_per_sheet:
                    break

        except Exception as e:
            self.logger.warning(f"Error comparing sheet {sheet_name}: {str(e)}")

        return changes

    def _compare_cell_pair(self, current_cell, comparison_cell, sheet_name: str, row: int, col: int) -> Optional[Dict[str, Any]]:
        """Compare a pair of cells and return change information if different"""
        current_value = current_cell.value
        comparison_value = comparison_cell.value

        # Skip if both cells are empty
        if current_value is None and comparison_value is None:
            return None

        # Determine change type
        change_type = self._determine_change_type(current_value, comparison_value)

        if change_type == "no_change":
            return None

        # Calculate change details
        change_info = {
            "sheet": sheet_name,
            "cell": f"{get_column_letter(col)}{row}",
            "row": row,
            "column": col,
            "old_value": comparison_value,
            "new_value": current_value,
            "change_type": change_type
        }

        # Add numeric change analysis
        if change_type in ["modified", "value_change"] and self._is_numeric(current_value) and self._is_numeric(comparison_value):
            change_info.update(self._analyze_numeric_change(current_value, comparison_value))

        # Check if this is a formula change
        change_info["formula_change"] = self._is_formula_change(current_cell, comparison_cell)

        # Flag major changes
        change_info["major_change"] = self._is_major_change(change_info)

        return change_info

    def _determine_change_type(self, current_value, comparison_value) -> str:
        """Determine the type of change between two cell values"""
        if current_value == comparison_value:
            return "no_change"

        if current_value is None:
            return "deleted"

        if comparison_value is None:
            return "added"

        # Both have values but different
        if self._is_numeric(current_value) and self._is_numeric(comparison_value):
            # Check tolerance for numeric values
            if abs(float(current_value) - float(comparison_value)) <= self.tolerance:
                return "no_change"
            return "value_change"

        return "modified"

    def _is_numeric(self, value) -> bool:
        """Check if a value is numeric"""
        if value is None:
            return False

        if isinstance(value, (int, float)):
            return True

        if isinstance(value, str):
            try:
                float(value.replace(',', '').replace('$', '').replace('%', ''))
                return True
            except ValueError:
                return False

        return False

    def _analyze_numeric_change(self, current_value, comparison_value) -> Dict[str, Any]:
        """Analyze numeric change details"""
        try:
            current_num = float(str(current_value).replace(',', '').replace('$', '').replace('%', ''))
            comparison_num = float(str(comparison_value).replace(',', '').replace('$', '').replace('%', ''))

            difference = current_num - comparison_num
            percent_change = (difference / comparison_num * 100) if comparison_num != 0 else 0

            return {
                "numeric_change": True,
                "difference": difference,
                "percent_change": round(percent_change, 2),
                "old_numeric": comparison_num,
                "new_numeric": current_num
            }

        except Exception:
            return {"numeric_change": False}

    def _is_formula_change(self, current_cell, comparison_cell) -> bool:
        """Check if the change involves formulas"""
        try:
            # This would require loading workbooks with formulas
            # For now, use data type as a proxy
            return (hasattr(current_cell, 'data_type') and current_cell.data_type == 'f') or \
                   (hasattr(comparison_cell, 'data_type') and comparison_cell.data_type == 'f')
        except Exception:
            return False

    def _is_major_change(self, change_info: Dict[str, Any]) -> bool:
        """Determine if a change is considered major"""
        # Major change criteria
        if change_info.get("numeric_change", False):
            percent_change = abs(change_info.get("percent_change", 0))
            return percent_change > 10  # >10% change is major

        if change_info.get("formula_change", False):
            return True

        # Large absolute changes
        if change_info.get("difference", 0) and abs(change_info["difference"]) > 1000000:  # >1M change
            return True

        return False

    def _compare_formulas(self) -> List[Dict[str, Any]]:
        """Compare formulas between workbooks"""
        # This would require loading workbooks with data_only=False
        # For now, return placeholder
        formula_changes = []

        try:
            # Load workbooks with formulas
            current_wb_formulas = openpyxl.load_workbook(str(self.current_path), data_only=False)
            comparison_wb_formulas = openpyxl.load_workbook(str(self.comparison_path), data_only=False)

            # Compare formulas in common sheets
            common_sheets = set(current_wb_formulas.sheetnames) & set(comparison_wb_formulas.sheetnames)
            common_sheets = [s for s in common_sheets if s not in self.ignore_sheets]

            for sheet_name in common_sheets:
                try:
                    current_sheet = current_wb_formulas[sheet_name]
                    comparison_sheet = comparison_wb_formulas[sheet_name]

                    # Sample check of formulas (limit for performance)
                    for row in range(1, min(current_sheet.max_row + 1, 100)):
                        for col in range(1, min(current_sheet.max_column + 1, 50)):
                            current_cell = current_sheet.cell(row=row, column=col)
                            comparison_cell = comparison_sheet.cell(row=row, column=col)

                            if (current_cell.data_type == 'f' or comparison_cell.data_type == 'f'):
                                if current_cell.value != comparison_cell.value:
                                    formula_changes.append({
                                        "sheet": sheet_name,
                                        "cell": f"{get_column_letter(col)}{row}",
                                        "old_formula": comparison_cell.value,
                                        "new_formula": current_cell.value
                                    })

                except Exception as e:
                    self.logger.warning(f"Error comparing formulas in sheet {sheet_name}: {str(e)}")

            current_wb_formulas.close()
            comparison_wb_formulas.close()

        except Exception as e:
            self.logger.warning(f"Error comparing formulas: {str(e)}")

        return formula_changes

    def _compare_metrics(self) -> Dict[str, Dict[str, Any]]:
        """Compare extracted metrics between versions"""
        metrics_comparison = {}

        try:
            # Extract metrics from both workbooks
            from metrics_extractor import MetricsExtractor

            current_extractor = MetricsExtractor(str(self.current_path), self.config)
            comparison_extractor = MetricsExtractor(str(self.comparison_path), self.config)

            current_metrics = current_extractor.extract()
            comparison_metrics = comparison_extractor.extract()

            # Compare each metric
            all_metrics = set(current_metrics.keys()) | set(comparison_metrics.keys())
            all_metrics.discard("_metadata")  # Skip metadata

            for metric_name in all_metrics:
                current_value = current_metrics.get(metric_name)
                comparison_value = comparison_metrics.get(metric_name)

                if current_value != comparison_value:
                    change_info = {
                        "old": comparison_value,
                        "new": current_value,
                        "changed": True
                    }

                    # Calculate percentage change for numeric values
                    if self._is_numeric(current_value) and self._is_numeric(comparison_value):
                        try:
                            current_num = float(current_value)
                            comparison_num = float(comparison_value)
                            difference = current_num - comparison_num
                            percent_change = (difference / comparison_num * 100) if comparison_num != 0 else 0

                            change_info.update({
                                "difference": difference,
                                "percent_change": round(percent_change, 2)
                            })
                        except Exception:
                            pass

                    metrics_comparison[metric_name] = change_info

        except Exception as e:
            self.logger.warning(f"Error comparing metrics: {str(e)}")

        return metrics_comparison

    def _generate_summary(self, comparison_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate summary of comparison results"""
        summary = {
            "total_cell_changes": len(comparison_results.get("cell_changes", [])),
            "major_changes": len([c for c in comparison_results.get("cell_changes", []) if c.get("major_change", False)]),
            "formula_changes": len(comparison_results.get("formula_changes", [])),
            "sheets_modified": len(set(c["sheet"] for c in comparison_results.get("cell_changes", []))),
            "metrics_changed": len(comparison_results.get("metrics_comparison", {})),
            "structural_changes": bool(
                comparison_results.get("structural_changes", {}).get("sheets_added") or
                comparison_results.get("structural_changes", {}).get("sheets_removed")
            )
        }

        # Add change distribution by type
        change_types = {}
        for change in comparison_results.get("cell_changes", []):
            change_type = change.get("change_type", "unknown")
            change_types[change_type] = change_types.get(change_type, 0) + 1

        summary["change_distribution"] = change_types

        return summary

    def _identify_alerts(self, comparison_results: Dict[str, Any]) -> List[Dict[str, str]]:
        """Identify significant changes that warrant user attention"""
        alerts = []

        # Check for formula errors
        for change in comparison_results.get("cell_changes", []):
            new_value = str(change.get("new_value", ""))
            if any(error in new_value for error in ["#REF!", "#VALUE!", "#N/A", "#NAME?"]):
                alerts.append({
                    "type": "ERROR",
                    "message": f"Formula error in {change['sheet']}!{change['cell']}: {new_value}",
                    "sheet": change["sheet"],
                    "cell": change["cell"]
                })

        # Check for major metric changes
        alert_threshold = self.config.get("alerts", {}).get("major_change_threshold", 0.10)
        for metric, change in comparison_results.get("metrics_comparison", {}).items():
            if change.get("percent_change", 0) and abs(change["percent_change"]) > alert_threshold * 100:
                alerts.append({
                    "type": "WARNING",
                    "message": f"{metric} changed by {change['percent_change']:.1f}%",
                    "metric": metric,
                    "change": change["percent_change"]
                })

        # Check for structural changes
        structural = comparison_results.get("structural_changes", {})
        if structural.get("sheets_added"):
            alerts.append({
                "type": "INFO",
                "message": f"Sheets added: {', '.join(structural['sheets_added'])}"
            })

        if structural.get("sheets_removed"):
            alerts.append({
                "type": "WARNING",
                "message": f"Sheets removed: {', '.join(structural['sheets_removed'])}"
            })

        return alerts

    def _close_workbooks(self):
        """Close workbooks to free memory"""
        try:
            if hasattr(self, 'current_wb'):
                self.current_wb.close()
            if hasattr(self, 'comparison_wb'):
                self.comparison_wb.close()
        except Exception as e:
            self.logger.warning(f"Error closing workbooks: {str(e)}")

    def export_comparison_report(self, output_path: str, comparison_results: Dict[str, Any]):
        """Export detailed comparison report to Excel"""
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = []
                for key, value in comparison_results.get("summary", {}).items():
                    summary_data.append({"Metric": key, "Value": value})

                if summary_data:
                    summary_df = pd.DataFrame(summary_data)
                    summary_df.to_excel(writer, sheet_name='Summary', index=False)

                # Cell changes
                if comparison_results.get("cell_changes"):
                    changes_df = pd.DataFrame(comparison_results["cell_changes"])
                    changes_df.to_excel(writer, sheet_name='Cell Changes', index=False)

                # Metrics comparison
                if comparison_results.get("metrics_comparison"):
                    metrics_data = []
                    for metric, change in comparison_results["metrics_comparison"].items():
                        metrics_data.append({
                            "Metric": metric,
                            "Old Value": change.get("old"),
                            "New Value": change.get("new"),
                            "Difference": change.get("difference"),
                            "Percent Change": change.get("percent_change")
                        })

                    if metrics_data:
                        metrics_df = pd.DataFrame(metrics_data)
                        metrics_df.to_excel(writer, sheet_name='Metrics Comparison', index=False)

                # Alerts
                if comparison_results.get("alerts"):
                    alerts_df = pd.DataFrame(comparison_results["alerts"])
                    alerts_df.to_excel(writer, sheet_name='Alerts', index=False)

            self.logger.info(f"Comparison report exported to {output_path}")

        except Exception as e:
            self.logger.error(f"Failed to export comparison report: {str(e)}")
            raise